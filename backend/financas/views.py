import datetime
import io
import unicodedata
from decimal import Decimal, InvalidOperation

import openpyxl
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django_filters import rest_framework as filters
from rest_framework import viewsets, permissions, status
from rest_framework.filters import SearchFilter
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Categoria, Transacao
from .serializers import CategoriaSerializer, TransacaoSerializer

MONTH_NAMES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3,
    'abril': 4, 'maio': 5, 'junho': 6,
    'julho': 7, 'agosto': 8, 'setembro': 9,
    'outubro': 10, 'novembro': 11, 'dezembro': 12,
}

IGNORED_SHEETS = {'gastos', 'acompanhamentos'}


def _normalize_cell(value):
    """Retorna o valor da célula em lowercase sem acentos."""
    if value is None:
        return ''
    text = unicodedata.normalize('NFD', str(value).strip())
    return ''.join(c for c in text if unicodedata.category(c) != 'Mn').lower()


def _normalize_header(value):
    text = _normalize_cell(value)
    return text.replace('.', '').replace(':', '').replace(' ', '')


def _is_value_header(value):
    return value.startswith('r$') or value in {'rs', 'r$'}


def _parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace('R$', '').replace('r$', '').replace(' ', '')
    if text.count(',') == 1 and text.count('.') >= 1:
        text = text.replace('.', '').replace(',', '.')
    else:
        text = text.replace(',', '.')
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_day(value):
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().day
    if isinstance(value, datetime.date):
        return value.day
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _find_header_row(rows):
    for idx, row in enumerate(rows):
        normalized = [_normalize_header(cell) for cell in row]
        if 'observacao' in normalized and 'dia' in normalized and any(_is_value_header(c) for c in normalized):
            return idx
    return None


def _get_category_name(category_row, start_col):
    if not category_row:
        return ''
    candidates = []
    for col in range(start_col, min(start_col + 3, len(category_row))):
        cell = category_row[col]
        if cell is not None and str(cell).strip():
            candidates.append(str(cell).strip())
    if candidates:
        return candidates[0]
    for col in range(max(0, start_col - 2), start_col):
        cell = category_row[col]
        if cell is not None and str(cell).strip():
            return str(cell).strip()
    return ''


class TransacaoFilter(filters.FilterSet):
    nome = filters.CharFilter(field_name='nome', lookup_expr='icontains')
    categoria_id = filters.NumberFilter(field_name='categoria__id')
    data_min = filters.DateFilter(field_name='data', lookup_expr='gte')
    data_max = filters.DateFilter(field_name='data', lookup_expr='lte')

    class Meta:
        model = Transacao
        fields = ['nome', 'categoria_id', 'data_min', 'data_max', 'tipo']


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [SearchFilter]
    search_fields = ['nome']


class TransacaoViewSet(viewsets.ModelViewSet):
    queryset = Transacao.objects.select_related('categoria').all()
    serializer_class = TransacaoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.DjangoFilterBackend, SearchFilter]
    filterset_class = TransacaoFilter
    search_fields = ['nome', 'descricao']

    def get_queryset(self):
        return self.queryset.filter(usuario=self.request.user)


class ImportXlsxView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'Arquivo XLSX é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)

        ano_param = request.data.get('ano')
        try:
            ano = int(ano_param) if ano_param else datetime.date.today().year
        except ValueError:
            return Response({'detail': 'Parâmetro "ano" inválido.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_obj.read()), data_only=True)
        except Exception as exc:
            return Response({'detail': f'Falha ao abrir XLSX: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

        total_created = 0
        all_errors = []

        for sheet_name in workbook.sheetnames:
            sheet_norm = _normalize_cell(sheet_name)

            if sheet_norm in IGNORED_SHEETS:
                continue

            mes_num = MONTH_NAMES_PT.get(sheet_norm)
            if mes_num is None:
                continue

            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            header_row_idx = _find_header_row(rows)
            if header_row_idx is None or header_row_idx == 0:
                all_errors.append({'aba': sheet_name, 'erro': 'Linha de cabecalho (Observacao/Dia/R$) nao encontrada.'})
                continue

            category_row = rows[header_row_idx - 1]
            header_norm = [_normalize_header(c) for c in rows[header_row_idx]]

            groups = []
            for col_idx, header in enumerate(header_norm):
                if header != 'observacao':
                    continue

                dia_col = next(
                    (j for j in range(col_idx + 1, min(col_idx + 4, len(header_norm))) if header_norm[j] == 'dia'),
                    None,
                )
                val_col = next(
                    (
                        j for j in range(col_idx + 1, min(col_idx + 4, len(header_norm)))
                        if _is_value_header(header_norm[j])
                    ),
                    None,
                )

                cat_name = _get_category_name(category_row, col_idx)
                if dia_col is not None and val_col is not None and cat_name:
                    groups.append({
                        'categoria': cat_name,
                        'obs_col': col_idx,
                        'dia_col': dia_col,
                        'val_col': val_col,
                    })

            if not groups:
                all_errors.append({'aba': sheet_name, 'erro': 'Nenhum grupo Observacao/Dia/R$ encontrado.'})
                continue

            created = 0
            for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
                for group in groups:
                    obs = ''
                    if group['obs_col'] < len(row):
                        obs = str(row[group['obs_col']] or '').strip()

                    if _normalize_cell(obs).startswith('total') or _normalize_cell(obs) in {'soma', 'saldo'}:
                        continue

                    dia = _parse_day(row[group['dia_col']] if group['dia_col'] < len(row) else None)
                    valor = _parse_decimal(row[group['val_col']] if group['val_col'] < len(row) else None)

                    if dia is None or valor is None:
                        continue

                    if valor <= 0:
                        continue

                    try:
                        data_transacao = datetime.date(ano, mes_num, dia)
                    except ValueError:
                        all_errors.append({
                            'aba': sheet_name,
                            'linha': row_idx,
                            'erro': f'Data invalida: dia {dia} em {sheet_name}/{ano}',
                        })
                        continue

                    cat_nome = group['categoria']
                    categoria = Categoria.objects.filter(nome__iexact=cat_nome, tipo=Categoria.DESPESA).first()
                    if not categoria:
                        categoria = Categoria.objects.create(nome=cat_nome, tipo=Categoria.DESPESA)

                    try:
                        Transacao.objects.create(
                            nome=obs or cat_nome,
                            descricao='',
                            valor=valor,
                            data=data_transacao,
                            tipo=Categoria.DESPESA,
                            categoria=categoria,
                            usuario=request.user,
                        )
                        created += 1
                    except Exception as exc:
                        all_errors.append({'aba': sheet_name, 'linha': row_idx, 'erro': str(exc)})

            total_created += created

        return Response({'created': total_created, 'errors': all_errors})


class RegisterView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username', '').strip()
        email = request.data.get('email', '').strip()
        password = request.data.get('password', '')
        password_confirm = request.data.get('password_confirm', '')

        if not username:
            return Response({'detail': 'O nome de usuário é obrigatório.'}, status=status.HTTP_400_BAD_REQUEST)

        if not password:
            return Response({'detail': 'A senha é obrigatória.'}, status=status.HTTP_400_BAD_REQUEST)

        if password != password_confirm:
            return Response({'detail': 'As senhas não coincidem.'}, status=status.HTTP_400_BAD_REQUEST)

        if User.objects.filter(username=username).exists():
            return Response({'detail': 'Nome de usuário já está em uso.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_password(password)
        except ValidationError as exc:
            return Response({'detail': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(username=username, email=email, password=password)
        return Response({'detail': 'Usuário criado com sucesso.', 'id': user.id, 'username': user.username}, status=status.HTTP_201_CREATED)

